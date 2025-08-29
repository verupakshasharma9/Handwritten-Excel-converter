from fastapi import FastAPI, APIRouter, File, UploadFile, HTTPException
from fastapi.responses import StreamingResponse
from fastapi.middleware.cors import CORSMiddleware
from dotenv import load_dotenv
from motor.motor_asyncio import AsyncIOMotorClient
import os
import logging
import io
import base64
import uuid
import json
from pathlib import Path
from pydantic import BaseModel, Field
from typing import List, Optional, Dict, Any
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, PatternFill, Alignment

# Try EmergentIntegrations
try:
    from emergentintegrations.llm.chat import LlmChat, UserMessage, ImageContent
    EMERGENT_AVAILABLE = True
except ImportError:
    EMERGENT_AVAILABLE = False
    print("⚠️  EmergentIntegrations not found. Install with:")
    print("pip install emergentintegrations --extra-index-url https://d33sy5i8bnduwe.cloudfront.net/simple/")

# Load environment variables
ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# MongoDB connection
mongo_url = os.environ.get('MONGO_URL', 'mongodb://localhost:27017')
db_name = os.environ.get('DB_NAME', 'handwritten_tables')
client = AsyncIOMotorClient(mongo_url)
db = client[db_name]

# Create FastAPI app
app = FastAPI(title="Handwritten Table Converter", version="1.0.0")
api_router = APIRouter(prefix="/api")

# ✅ FIXED CORS Middleware (allow all during dev)
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],  # for dev, allow all origins
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# -------------------------
# Models
# -------------------------
class TableData(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    filename: str
    extracted_data: List[List[str]]
    created_at: datetime = Field(default_factory=datetime.utcnow)

class ProcessingResult(BaseModel):
    success: bool
    message: str
    table_data: Optional[List[List[str]]] = None
    processing_id: Optional[str] = None

# -------------------------
# Utility Functions
# -------------------------
def image_to_base64(image_bytes: bytes) -> str:
    return base64.b64encode(image_bytes).decode('utf-8')

async def extract_table_from_image(image_bytes: bytes, filename: str) -> Dict[str, Any]:
    try:
        if not EMERGENT_AVAILABLE:
            return {
                "success": True,
                "table_data": [
                    ["Name", "Age", "City"],
                    ["John", "25", "NYC"],
                    ["Alice", "30", "LA"]
                ],
                "message": "Table extracted successfully (Mock data)"
            }

        image_base64 = image_to_base64(image_bytes)
        chat = LlmChat(
            api_key=os.environ.get('EMERGENT_LLM_KEY'),
            session_id=f"table_extraction_{uuid.uuid4()}",
            system_message="You are an expert at analyzing handwritten tables and extracting structured data."
        ).with_model("openai", "gpt-4o")

        image_content = ImageContent(image_base64=image_base64)

        prompt = """Analyze this handwritten table image and extract all data into structured format.

Return ONLY valid JSON array like:
[
  ["Header1", "Header2"],
  ["Row1Col1", "Row1Col2"]
]"""

        user_message = UserMessage(text=prompt, file_contents=[image_content])
        response = await chat.send_message(user_message)

        try:
            response_text = response.strip()
            if response_text.startswith('```'):
                response_text = response_text.strip('`json').strip('`')
            table_data = json.loads(response_text.strip())

            if not isinstance(table_data, list) or not table_data:
                raise ValueError("Invalid table data format")

            return {
                "success": True,
                "table_data": table_data,
                "message": "Table extracted successfully"
            }
        except json.JSONDecodeError as e:
            logging.error(f"JSON parsing error: {e}")
            return {"success": False, "message": "Failed to parse extracted data", "table_data": None}

    except Exception as e:
        logging.error(f"OCR processing error: {e}")
        return {"success": False, "message": f"Error: {str(e)}", "table_data": None}

def create_excel_file(table_data: List[List[str]], filename: str) -> io.BytesIO:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Extracted Table"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    border = Border(left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin'))
    center_alignment = Alignment(horizontal='center', vertical='center')

    for row_idx, row_data in enumerate(table_data, 1):
        for col_idx, cell_value in enumerate(row_data, 1):
            cell = worksheet.cell(row=row_idx, column=col_idx, value=cell_value)
            cell.border = border
            cell.alignment = center_alignment
            if row_idx == 1:
                cell.font = header_font
                cell.fill = header_fill

    for column in worksheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        worksheet.column_dimensions[column_letter].width = min(max_length + 2, 50)

    excel_buffer = io.BytesIO()
    workbook.save(excel_buffer)
    excel_buffer.seek(0)
    return excel_buffer

# -------------------------
# API Routes
# -------------------------
@api_router.get("/")
async def root():
    return {"message": "Handwritten Table Converter API", "status": "running"}

@api_router.post("/upload-image", response_model=ProcessingResult)
async def upload_image(file: UploadFile = File(...)):
    if not file.content_type.startswith('image/'):
        raise HTTPException(status_code=400, detail="Only image files are allowed")

    image_bytes = await file.read()
    result = await extract_table_from_image(image_bytes, file.filename)

    if result["success"]:
        table_record = TableData(filename=file.filename, extracted_data=result["table_data"])
        await db.table_extractions.insert_one(table_record.dict())
        return ProcessingResult(success=True, message=result["message"],
                                table_data=result["table_data"], processing_id=table_record.id)
    else:
        return ProcessingResult(success=False, message=result["message"])

@api_router.post("/generate-excel/{processing_id}")
async def generate_excel(processing_id: str):
    record = await db.table_extractions.find_one({"id": processing_id})
    if not record:
        raise HTTPException(status_code=404, detail="Processing record not found")

    excel_buffer = create_excel_file(record["extracted_data"], record["filename"])
    excel_filename = f"{record['filename'].rsplit('.',1)[0]}_extracted.xlsx"

    return StreamingResponse(
        io.BytesIO(excel_buffer.read()),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={excel_filename}"}
    )

@api_router.get("/extractions", response_model=List[TableData])
async def get_extractions():
    records = await db.table_extractions.find().sort("created_at", -1).to_list(50)
    return [TableData(**record) for record in records]

# Include router
app.include_router(api_router)

# Logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')

@app.on_event("startup")
async def startup_event():
    logging.info("🚀 API Started")
    logging.info(f"📊 Database: {mongo_url}/{db_name}")
    logging.info("🧠 AI Integration: " + ("EmergentIntegrations" if EMERGENT_AVAILABLE else "Mock mode"))

@app.on_event("shutdown")
async def shutdown_event():
    client.close()
    logging.info("📴 Application shutdown complete")

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)
